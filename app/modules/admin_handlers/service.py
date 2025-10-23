import os
from datetime import datetime
from io import BytesIO

from fastapi import Depends
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font

from typing import AsyncIterator

from starlette.datastructures import UploadFile

from app.core.config import settings
from app.core.db.session import get_session
from app.modules.admin_handlers.enums.RecipientsEnum import RecipientsEnum
from app.modules.admin_handlers.schemas.api.BroadcastingSchema import BroadcastingSchema
from app.modules.goods.entities import GoodEntity, GoodVariationEntity
from app.modules.orders.entities import OrderEntity, OrderDetailsEntity
from app.modules.payments.entities import PaymentEntity
from app.modules.payments.enums.payment_statuses import PaymentStatuses
from app.modules.users.entities import UserEntity

from aiogram import Bot
from aiogram.types import URLInputFile

bot = Bot(token=settings.BOT_TOKEN)


class InvalidUsers(ValueError):
    pass


class SendingMessages:
    def __init__(self, db: AsyncSession = Depends(get_session)):
        self.db = db

    async def send_message(
        self,
        data: BroadcastingSchema,
    ):
        """
        Отправка сообщений пользователям

        Args:
            data: BroadcastingSchema - данные для рассылки
        """

        if RecipientsEnum.ALL in data.recipients:
            stmt = select(UserEntity.telegram_id).where(
                UserEntity.telegram_id.is_not(None)
            )
        elif RecipientsEnum.ADMINS in data.recipients:
            stmt = select(UserEntity.telegram_id).where(
                UserEntity.is_admin == True, UserEntity.telegram_id.is_not(None)
            )
        else:
            raise InvalidUsers("No valid recipients specified")

        result = await self.db.execute(stmt)
        user_chat_ids = [row[0] for row in result.all()]

        media = None
        if data.photo:
            media = URLInputFile(data.photo)

        for chat_id in user_chat_ids:
            (
                await bot.send_photo(
                    chat_id=chat_id,
                    photo=media,
                    caption=data.message,
                    parse_mode="HTML",
                )
                if media
                else await bot.send_message(
                    chat_id=chat_id, text=data.message, parse_mode="HTML"
                )
            )


class ExcelService:
    def __init__(self, db: AsyncSession = Depends(get_session)):
        self.db = db

    async def create_excel(self, start_date: datetime, end_date: datetime):
        """
        Генерация Excel файла с заказами за указанный период

        Args:
            start_date: Начальная дата периода
            end_date: Конечная дата периода

        Returns:
            Поток байтов с содержимым Excel файла
        """

        # Получаем данные из БД
        stmt = (
            select(
                OrderEntity.id,
                UserEntity.first_name,
                UserEntity.last_name,
                UserEntity.phone,
                UserEntity.email,
                GoodEntity.id.label("good_id"),
                GoodEntity.title.label("good_name"),
                GoodVariationEntity.id.label("variation_id"),
                GoodVariationEntity.title.label("variation_title"),
                OrderDetailsEntity.price,
                OrderDetailsEntity.quantity,
                (OrderDetailsEntity.price * OrderDetailsEntity.quantity).label(
                    "amount"
                ),
            )
            .join(OrderEntity.user)
            .join(OrderEntity.details)
            .join(OrderDetailsEntity.variation)
            .join(GoodVariationEntity.good)
            .join(OrderEntity.payments)
            .where(OrderEntity.created_at >= start_date)
            .where(OrderEntity.created_at <= end_date)
            .where(PaymentEntity.status == PaymentStatuses.SUCCESS)
            .order_by(OrderEntity.created_at)
        )

        result = await self.db.execute(stmt)
        orders_data = result.all()

        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"

        headers = [
            "ID заказа",
            "Имя",
            "Фамилия",
            "Телефон",
            "Email",
            "ID товара",
            "Наименование товара",
            "ID вариации",
            "Наименование вариации",
            "Цена",
            "Количество",
            "Сумма",
        ]

        ws.append(headers)

        for cell in ws[1]:
            cell.font = Font(bold=True)

        for order in orders_data:
            ws.append(
                [
                    order.id,
                    order.first_name,
                    order.last_name,
                    order.phone,
                    order.email,
                    order.good_id,
                    order.good_name,
                    order.variation_id,
                    order.variation_title,
                    order.price,
                    order.quantity,
                    order.amount,
                ]
            )

        total_amount = sum(order.amount for order in orders_data)
        ws.append(["ИТОГО:", "", "", "", "", "", "", "", "", "", "", total_amount])
        ws[f"L{ws.max_row}"].font = Font(bold=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return output
